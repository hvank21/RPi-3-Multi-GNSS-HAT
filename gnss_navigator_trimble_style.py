#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GNSS Navigator Trimble-style 720x480 для Raspberry Pi + L76K GPS HAT.

Що збережено:
- NAV / SKY / SNR / DIAG / SATS / MAP
- веб-мапи Online/Offline через локальний HTTP server
- фільтр сузір'я у SATS
- запис SATS у Excel з вибором сузір'я для запису

Що додано:
- вкладка "Сузір'я" для перемикання L76K: GPS / BeiDou / GLONASS та комбіновані режими
- відправка PCAS-команд у той самий UART, з якого читається NMEA

Запуск:
    python3 gnss_navigator_trimble_style.py

UART:
    GNSS_SERIAL_PORT=/dev/serial0 python3 gnss_navigator_trimble_style.py
    GNSS_SERIAL_PORT=/dev/ttyUSB0  python3 gnss_navigator_trimble_style.py
"""

import json
import math
import os
import queue
import threading
import time
import webbrowser
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Dict, Optional
from urllib.parse import unquote

import serial
import tkinter as tk
from tkinter import ttk, messagebox

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


# ============================================================
# Налаштування
# ============================================================
SERIAL_PORT = os.environ.get("GNSS_SERIAL_PORT", "/dev/serial0")
BAUDRATE = int(os.environ.get("GNSS_BAUDRATE", "9600"))
SERIAL_TIMEOUT = 1.0

REFRESH_MS = 400
WINDOW_WIDTH = 720
WINDOW_HEIGHT = 480
FULLSCREEN = False
TITLE = "GNSS Navigator"

HEADER_FONT = ("DejaVu Sans", 14, "bold")
STATUS_FONT = ("DejaVu Sans", 8)
TAB_FONT = ("DejaVu Sans", 9, "bold")
NAV_TITLE_FONT = ("DejaVu Sans", 9, "bold")
NAV_VALUE_FONT = ("DejaVu Sans", 16, "bold")
SECTION_FONT = ("DejaVu Sans", 11, "bold")
MONO_FONT = ("DejaVu Sans Mono", 9)
SMALL_FONT = ("DejaVu Sans", 8)
SMALL_BOLD_FONT = ("DejaVu Sans", 8, "bold")

MAP_ONLINE_PORT = int(os.environ.get("GNSS_MAP_ONLINE_PORT", "8080"))
MAP_OFFLINE_PORT = int(os.environ.get("GNSS_MAP_OFFLINE_PORT", "8081"))
TRACK_MAX_POINTS = 3000
TILES_ROOT = str(Path.home() / "gnss_tiles")
LOGS_ROOT = Path.home() / "gnss_logs"
START_LAT = 50.4501
START_LON = 30.5234
START_ZOOM = 15
SATS_LOG_MIN_INTERVAL_SEC = 1.0
TRACK_MIN_INTERVAL_SEC = 1.0

BG = "#101318"
PANEL = "#171c24"
PANEL2 = "#243140"
TEXT = "#edf3fb"
TEXT2 = "#dbe7f6"
MUTED = "#a9b6c8"
GREEN = "#6cc46e"
YELLOW = "#d6c24a"
ORANGE = "#e39b3a"
RED = "#d95763"
BLUE = "#2d4258"


# ============================================================
# Дані GNSS
# ============================================================
@dataclass
class Satellite:
    prn: str
    elev: Optional[int] = None
    azim: Optional[int] = None
    snr: Optional[int] = None
    talker: str = ""
    used: bool = False
    last_seen: float = field(default_factory=time.time)


@dataclass
class GNSSState:
    lat: Optional[float] = None
    lon: Optional[float] = None
    alt_m: Optional[float] = None
    speed_knots: Optional[float] = None
    course_deg: Optional[float] = None
    utc_time: Optional[str] = None
    utc_date: Optional[str] = None
    fix_quality: int = 0
    fix_type: int = 1
    sats_used: int = 0              # USED, перераховано тільки з GSA
    gga_sats_used: int = 0          # сире поле GGA[7], лише для діагностики
    used_by_talker: Dict[str, set] = field(default_factory=dict)
    pdop: Optional[float] = None
    hdop: Optional[float] = None
    vdop: Optional[float] = None
    mode_2d3d: Optional[str] = None
    last_sentence: str = ""
    last_update: float = 0.0
    satellites: Dict[str, Satellite] = field(default_factory=dict)
    track: list = field(default_factory=list)
    last_track_time: float = 0.0
    constellation_mode: str = "Не вибрано"
    last_cmd: str = ""
    last_cmd_status: str = ""


state_lock = threading.Lock()
gnss_state = GNSSState()


# ============================================================
# Допоміжні функції
# ============================================================
def int_safe(v, default=0):
    try:
        if v is None or v == "":
            return default
        return int(v)
    except Exception:
        return default


def float_safe(v, default=None):
    try:
        if v is None or v == "":
            return default
        return float(str(v).split("*", 1)[0])
    except Exception:
        return default


def parse_nmea_lat(raw: str, hemi: str) -> Optional[float]:
    if not raw or not hemi:
        return None
    try:
        deg = int(raw[:2])
        minutes = float(raw[2:])
        val = deg + minutes / 60.0
        return -val if hemi.upper() == "S" else val
    except Exception:
        return None


def parse_nmea_lon(raw: str, hemi: str) -> Optional[float]:
    if not raw or not hemi:
        return None
    try:
        deg = int(raw[:3])
        minutes = float(raw[3:])
        val = deg + minutes / 60.0
        return -val if hemi.upper() == "W" else val
    except Exception:
        return None


def knots_to_kmh(knots: Optional[float]) -> Optional[float]:
    if knots is None:
        return None
    return knots * 1.852


def fmt_float(v: Optional[float], digits: int = 2, suffix: str = "") -> str:
    if v is None:
        return "--"
    return f"{v:.{digits}f}{suffix}"


def safe_int_sort(v):
    try:
        return int(v)
    except Exception:
        return 9999


def snr_to_color(snr: Optional[int]) -> str:
    if snr is None:
        return "#4e657d"
    if snr < 20:
        return RED
    if snr < 30:
        return ORANGE
    if snr < 40:
        return YELLOW
    return GREEN


def talker_name(talker: str) -> str:
    return {
        "GP": "GPS",
        "GL": "GLONASS",
        "GA": "Galileo",
        "GB": "BeiDou",
        "BD": "BeiDou",
        "GN": "Mixed",
        "QZ": "QZSS",
    }.get(talker, talker or "--")


SATS_FILTER_VALUES = [
    "ALL",
    "GPS",
    "QZSS",
    "BeiDou",
    "GLONASS",
    "GPS+BeiDou",
    "GPS+GLONASS",
    "BeiDou+GLONASS",
    "GPS+BeiDou+GLONASS",
    "Galileo",
    "Mixed",
]


def filter_to_talkers(filter_name: str):
    f = (filter_name or "ALL").strip()
    # QZSS працює на L1 і часто логічно йде разом з GPS, тому в режимі GPS
    # показуємо GP + QZ. Окремий фільтр QZSS теж залишений.
    if f == "GPS":
        return {"GP", "QZ"}
    if f == "QZSS":
        return {"QZ"}
    if f == "BeiDou":
        return {"GB", "BD"}
    if f == "GLONASS":
        return {"GL"}
    if f == "GPS+BeiDou":
        return {"GP", "QZ", "GB", "BD", "GN"}
    if f == "GPS+GLONASS":
        return {"GP", "QZ", "GL", "GN"}
    if f == "BeiDou+GLONASS":
        return {"GB", "BD", "GL", "GN"}
    if f == "GPS+BeiDou+GLONASS":
        return {"GP", "QZ", "GB", "BD", "GL", "GN"}
    if f == "Galileo":
        return {"GA"}
    if f == "Mixed":
        return {"GN"}
    return None


def mode_to_sats_filter(mode_name: str) -> Optional[str]:
    """
    Режим L76K -> фільтр таблиці SATS.
    Restart не змінює фільтр, бо це не вибір сузір'я.
    """
    m = (mode_name or "").strip()
    if m in (
        "GPS",
        "BeiDou",
        "GLONASS",
        "GPS+BeiDou",
        "GPS+GLONASS",
        "BeiDou+GLONASS",
        "GPS+BeiDou+GLONASS",
    ):
        return m
    return None


def satellite_allowed_by_filter(sat: Satellite, filter_name: str) -> bool:
    allowed = filter_to_talkers(filter_name)
    if allowed is None:
        return True
    return sat.talker in allowed


def recompute_used_flags_locked():
    """
    USED рахуємо тільки з GSA.
    GSV = видимі супутники, але не USED.
    GGA[7] зберігаємо окремо як сире поле, бо воно може включати всі системи.
    """
    total = 0
    for sat in gnss_state.satellites.values():
        used = False
        # Нормальні рядки: GPGSA/BDGSA/GBGSA/GLGSA/QZGSA
        if sat.prn in gnss_state.used_by_talker.get(sat.talker, set()):
            used = True
        # Деякі модулі дають GNGSA як mixed. Тоді PRN зіставляємо з видимими GSV.
        if sat.prn in gnss_state.used_by_talker.get("GN", set()):
            used = True
        sat.used = used
        if used:
            total += 1
    gnss_state.sats_used = total


def count_visible_used(sats, filter_name: str = "ALL"):
    filtered = [sat for sat in sats if satellite_allowed_by_filter(sat, filter_name)]
    return len(filtered), sum(1 for sat in filtered if sat.used)


def clear_satellite_cache():
    with state_lock:
        gnss_state.satellites.clear()
        gnss_state.used_by_talker.clear()
        gnss_state.sats_used = 0
        gnss_state.gga_sats_used = 0


def purge_satellites_for_filter(filter_name: str):
    """
    При перемиканні режиму прибираємо старі GSV-записи інших систем,
    щоб у SATS не висіли супутники попереднього режиму ще 20 секунд.
    """
    allowed = filter_to_talkers(filter_name)
    if allowed is None:
        return
    with state_lock:
        gnss_state.satellites = {
            key: sat for key, sat in gnss_state.satellites.items()
            if sat.talker in allowed
        }
        for talker in list(gnss_state.used_by_talker.keys()):
            if talker != "GN" and talker not in allowed:
                del gnss_state.used_by_talker[talker]
        recompute_used_flags_locked()


def format_utc(t: Optional[str], d: Optional[str]) -> str:
    if not t:
        return "--:--:--"
    ts = t[:6]
    if len(ts) == 6:
        ts = f"{ts[0:2]}:{ts[2:4]}:{ts[4:6]}"
    if d and len(d) == 6:
        return f"{ts} | {d[0:2]}.{d[2:4]}.20{d[4:6]}"
    return ts


def fix_text(st: GNSSState) -> str:
    if st.fix_quality <= 0:
        return "NO FIX"
    if st.fix_type == 3:
        return "3D FIX"
    if st.fix_type == 2:
        return "2D FIX"
    return "FIX"


def distance_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2.0) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2.0) ** 2
    return 2.0 * r * math.asin(math.sqrt(a))


def nmea_checksum(payload: str) -> str:
    c = 0
    for ch in payload:
        c ^= ord(ch)
    return f"{c:02X}"


def make_pcas(payload: str) -> str:
    return f"${payload}*{nmea_checksum(payload)}"


def update_track(lat: float, lon: float):
    with state_lock:
        now = time.time()
        tr = gnss_state.track
        if not tr:
            tr.append([lat, lon])
            gnss_state.last_track_time = now
            return
        last_lat, last_lon = tr[-1]
        if abs(last_lat - lat) < 1e-9 and abs(last_lon - lon) < 1e-9:
            return
        if (now - gnss_state.last_track_time) >= TRACK_MIN_INTERVAL_SEC:
            tr.append([lat, lon])
            gnss_state.last_track_time = now
            if len(tr) > TRACK_MAX_POINTS:
                del tr[:len(tr) - TRACK_MAX_POINTS]


# ============================================================
# NMEA парсер
# ============================================================
class NMEAParser:
    def parse_line(self, raw_line: str):
        line = raw_line.strip()
        if not line.startswith("$"):
            return
        body = line[1:].split("*", 1)[0]
        parts = body.split(",")
        if not parts or len(parts[0]) < 5:
            return
        kind = parts[0][2:]
        talker = parts[0][:2]
        with state_lock:
            gnss_state.last_sentence = line
            gnss_state.last_update = time.time()
        try:
            if kind == "GGA":
                self._parse_gga(parts)
            elif kind == "RMC":
                self._parse_rmc(parts)
            elif kind == "GSA":
                self._parse_gsa(parts, talker)
            elif kind == "GSV":
                self._parse_gsv(parts, talker)
        except Exception:
            pass

    def _parse_gga(self, parts):
        lat = parse_nmea_lat(parts[2], parts[3]) if len(parts) > 4 else None
        lon = parse_nmea_lon(parts[4], parts[5]) if len(parts) > 6 else None
        with state_lock:
            gnss_state.utc_time = parts[1] if len(parts) > 1 and parts[1] else gnss_state.utc_time
            gnss_state.lat = lat if lat is not None else gnss_state.lat
            gnss_state.lon = lon if lon is not None else gnss_state.lon
            gnss_state.fix_quality = int_safe(parts[6], gnss_state.fix_quality) if len(parts) > 6 else gnss_state.fix_quality
            gnss_state.gga_sats_used = int_safe(parts[7], gnss_state.gga_sats_used) if len(parts) > 7 else gnss_state.gga_sats_used
            gnss_state.hdop = float_safe(parts[8], gnss_state.hdop) if len(parts) > 8 else gnss_state.hdop
            gnss_state.alt_m = float_safe(parts[9], gnss_state.alt_m) if len(parts) > 9 else gnss_state.alt_m
        if lat is not None and lon is not None:
            update_track(lat, lon)

    def _parse_rmc(self, parts):
        lat = parse_nmea_lat(parts[3], parts[4]) if len(parts) > 5 else None
        lon = parse_nmea_lon(parts[5], parts[6]) if len(parts) > 7 else None
        with state_lock:
            gnss_state.utc_time = parts[1] if len(parts) > 1 and parts[1] else gnss_state.utc_time
            gnss_state.lat = lat if lat is not None else gnss_state.lat
            gnss_state.lon = lon if lon is not None else gnss_state.lon
            gnss_state.speed_knots = float_safe(parts[7], gnss_state.speed_knots) if len(parts) > 7 else gnss_state.speed_knots
            gnss_state.course_deg = float_safe(parts[8], gnss_state.course_deg) if len(parts) > 8 else gnss_state.course_deg
            gnss_state.utc_date = parts[9] if len(parts) > 9 and parts[9] else gnss_state.utc_date
        if lat is not None and lon is not None:
            update_track(lat, lon)

    def _parse_gsa(self, parts, talker):
        # GSA містить PRN супутників, реально використаних у розв'язку.
        # Саме GSA, а не GSV і не GGA[7], є джерелом прапорця USED у SATS/NAV.
        used_prns = set()
        for idx in range(3, 15):
            if idx < len(parts) and parts[idx]:
                prn = parts[idx].split("*", 1)[0].strip()
                if prn:
                    used_prns.add(prn)
        with state_lock:
            if len(parts) > 2:
                gnss_state.mode_2d3d = parts[2] if parts[2] else gnss_state.mode_2d3d
                gnss_state.fix_type = int_safe(parts[2], gnss_state.fix_type)
            if len(parts) > 15:
                gnss_state.pdop = float_safe(parts[15], gnss_state.pdop)
            if len(parts) > 16:
                gnss_state.hdop = float_safe(parts[16], gnss_state.hdop)
            if len(parts) > 17:
                gnss_state.vdop = float_safe(parts[17], gnss_state.vdop)
            gnss_state.used_by_talker[talker] = used_prns
            recompute_used_flags_locked()

    def _parse_gsv(self, parts, talker):
        idx = 4
        while idx + 3 < len(parts):
            prn = parts[idx].strip()
            if prn:
                elev = int_safe(parts[idx + 1], None)
                azim = int_safe(parts[idx + 2], None)
                snr_field = parts[idx + 3].split("*", 1)[0]
                snr = int_safe(snr_field, None)
                key = f"{talker}-{prn}"
                with state_lock:
                    sat = gnss_state.satellites.get(key, Satellite(prn=prn, talker=talker))
                    sat.elev = elev
                    sat.azim = azim
                    sat.snr = snr
                    sat.talker = talker
                    sat.last_seen = time.time()
                    gnss_state.satellites[key] = sat
                    recompute_used_flags_locked()
            idx += 4


# ============================================================
# Читання UART + запис PCAS-команд у той самий відкритий порт
# ============================================================
class GNSSReader(threading.Thread):
    def __init__(self, parser: NMEAParser, port: str, baudrate: int):
        super().__init__(daemon=True)
        self.parser = parser
        self.port = port
        self.baudrate = baudrate
        self.last_error = ""
        self._stop_event = threading.Event()
        self.ser = None
        self.ser_lock = threading.Lock()
        self.cmd_queue = queue.Queue()

    def stop(self):
        self._stop_event.set()

    @staticmethod
    def normalize_cmd(cmd: str) -> Optional[str]:
        cmd = (cmd or "").strip()
        if not cmd.startswith("$") or "*" not in cmd:
            return None
        return cmd

    def queue_command(self, cmd: str, delay_after_sec: float = 0.25) -> bool:
        """
        Ставить PCAS-команду в чергу потоку UART.
        Це надійніше, ніж писати з GUI-потоку під час blocking readline().
        """
        cmd = self.normalize_cmd(cmd)
        if cmd is None:
            self.last_error = f"Некоректна команда: {cmd}"
            return False
        self.cmd_queue.put((cmd, delay_after_sec))
        self.last_error = ""
        return True

    def send_command(self, cmd: str) -> bool:
        # Сумісність зі старими викликами: тепер тільки черга.
        return self.queue_command(cmd)

    def queue_commands(self, commands) -> bool:
        ok = True
        for cmd in commands:
            ok = self.queue_command(cmd) and ok
        return ok

    def _drain_command_queue(self, ser):
        while True:
            try:
                cmd, delay_after_sec = self.cmd_queue.get_nowait()
            except queue.Empty:
                break
            payload = (cmd + "\r\n").encode("ascii", errors="strict")
            try:
                ser.write(payload)
                ser.flush()
                self.last_error = ""
                time.sleep(delay_after_sec)
            except Exception as e:
                self.last_error = f"TX error {cmd}: {e}"

    def run(self):
        while not self._stop_event.is_set():
            try:
                with serial.Serial(self.port, self.baudrate, timeout=0.20, write_timeout=1.0) as ser:
                    with self.ser_lock:
                        self.ser = ser
                    self.last_error = ""
                    while not self._stop_event.is_set():
                        self._drain_command_queue(ser)
                        raw = ser.readline()
                        self._drain_command_queue(ser)
                        if not raw:
                            continue
                        line = raw.decode("ascii", errors="ignore")
                        if line:
                            self.parser.parse_line(line)
            except Exception as e:
                self.last_error = str(e)
                time.sleep(1.0)
            finally:
                with self.ser_lock:
                    self.ser = None


# ============================================================
# HTTP карта: online + offline
# ============================================================
ONLINE_HTML = '''<!DOCTYPE html>
<html lang="uk">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GNSS Map Online</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
html, body { margin:0; padding:0; height:100%; background:#0f141a; color:#eef4fb; font-family:Arial,sans-serif; }
#map { position:absolute; top:0; left:0; right:0; bottom:0; }
#panel { position:absolute; top:10px; left:10px; z-index:1000; background:rgba(20,27,36,0.92); color:#eef4fb; border-radius:12px; padding:12px 14px; min-width:280px; box-shadow:0 6px 24px rgba(0,0,0,0.35); }
#panel h2 { margin:0 0 10px 0; font-size:20px; }
.row { margin:5px 0; font-size:14px; }
.label { color:#9db1c8; display:inline-block; min-width:92px; }
.small { font-size:12px; color:#a8bacd; margin-top:8px; }
</style>
</head>
<body>
<div id="map"></div>
<div id="panel">
<h2>GNSS Map Online</h2>
<div class="row"><span class="label">Fix:</span><span id="fix">--</span></div>
<div class="row"><span class="label">Lat:</span><span id="lat">--</span></div>
<div class="row"><span class="label">Lon:</span><span id="lon">--</span></div>
<div class="row"><span class="label">Alt:</span><span id="alt">--</span></div>
<div class="row"><span class="label">Speed:</span><span id="speed">--</span></div>
<div class="row"><span class="label">Course:</span><span id="course">--</span></div>
<div class="row"><span class="label">Sats:</span><span id="sats">--</span></div>
<div class="row"><span class="label">HDOP:</span><span id="hdop">--</span></div>
<div class="row"><span class="label">UTC:</span><span id="utc">--</span></div>
<div class="small">OpenStreetMap online tiles. Потрібен інтернет.</div>
</div>
<script>
const map = L.map('map').setView([50.4501, 30.5234], 15);
L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', { maxZoom:19, attribution:'&copy; OpenStreetMap contributors' }).addTo(map);
const marker = L.marker([50.4501, 30.5234]).addTo(map);
const track = L.polyline([], {weight:4}).addTo(map);
let firstFix = true;
function fmt(v, digits=6, suffix='') { if(v===null||v===undefined) return '--'; if(typeof v==='number') return v.toFixed(digits)+suffix; return String(v)+suffix; }
function knotsToKmh(k){ if(k===null||k===undefined) return null; return k*1.852; }
function fixText(d){ if((d.fix_quality||0)<=0) return 'NO FIX'; return d.fix_type===3 ? '3D FIX' : (d.fix_type===2 ? '2D FIX' : 'FIX'); }
async function upd(){
 const r = await fetch('/api/position'); const d = await r.json();
 document.getElementById('fix').textContent = fixText(d);
 document.getElementById('lat').textContent = fmt(d.lat,6);
 document.getElementById('lon').textContent = fmt(d.lon,6);
 document.getElementById('alt').textContent = d.alt_m==null ? '--' : d.alt_m.toFixed(1)+' m';
 const sp = knotsToKmh(d.speed_knots);
 document.getElementById('speed').textContent = sp==null ? '--' : sp.toFixed(1)+' km/h';
 document.getElementById('course').textContent = d.course_deg==null ? '--' : d.course_deg.toFixed(1)+'°';
 document.getElementById('sats').textContent = d.sats_used ?? '--';
 document.getElementById('hdop').textContent = d.hdop==null ? '--' : d.hdop.toFixed(2);
 document.getElementById('utc').textContent = d.utc_text || '--';
 if(d.lat!=null && d.lon!=null){ marker.setLatLng([d.lat,d.lon]); if(Array.isArray(d.track)) track.setLatLngs(d.track); if(firstFix){ map.setView([d.lat,d.lon],17); firstFix=false; } }
}
upd(); setInterval(upd, 1000);
</script>
</body>
</html>'''


def offline_html() -> str:
    return f'''<!DOCTYPE html>
<html lang="uk">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GNSS Map Offline</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
html, body {{ margin:0; padding:0; height:100%; background:#0f141a; color:#eef4fb; font-family:Arial,sans-serif; }}
#map {{ position:absolute; top:0; left:0; right:0; bottom:0; }}
#panel {{ position:absolute; top:10px; left:10px; z-index:1000; background:rgba(20,27,36,0.92); color:#eef4fb; border-radius:12px; padding:12px 14px; min-width:300px; box-shadow:0 6px 24px rgba(0,0,0,0.35); }}
#panel h2 {{ margin:0 0 10px 0; font-size:20px; }}
.row {{ margin:5px 0; font-size:14px; }}
.label {{ color:#9db1c8; display:inline-block; min-width:96px; }}
.small {{ font-size:12px; color:#a8bacd; margin-top:8px; line-height:1.4; }}
</style>
</head>
<body>
<div id="map"></div>
<div id="panel">
<h2>GNSS Map Offline</h2>
<div class="row"><span class="label">Fix:</span><span id="fix">--</span></div>
<div class="row"><span class="label">Lat:</span><span id="lat">--</span></div>
<div class="row"><span class="label">Lon:</span><span id="lon">--</span></div>
<div class="row"><span class="label">Alt:</span><span id="alt">--</span></div>
<div class="row"><span class="label">Speed:</span><span id="speed">--</span></div>
<div class="row"><span class="label">Course:</span><span id="course">--</span></div>
<div class="row"><span class="label">Sats:</span><span id="sats">--</span></div>
<div class="row"><span class="label">HDOP:</span><span id="hdop">--</span></div>
<div class="row"><span class="label">UTC:</span><span id="utc">--</span></div>
<div class="small">Локальні тайли: <b>{TILES_ROOT}</b><br>Формат: <b>/tiles/z/x/y.png</b></div>
</div>
<script>
const map = L.map('map').setView([{START_LAT}, {START_LON}], {START_ZOOM});
L.tileLayer('/tiles/{{z}}/{{x}}/{{y}}.png', {{ maxZoom:19, attribution:'Offline tiles' }}).addTo(map);
const marker = L.marker([{START_LAT}, {START_LON}]).addTo(map);
const track = L.polyline([], {{weight:4}}).addTo(map);
let firstFix = true;
function fmt(v, digits=6, suffix='') {{ if(v===null||v===undefined) return '--'; if(typeof v==='number') return v.toFixed(digits)+suffix; return String(v)+suffix; }}
function knotsToKmh(k){{ if(k===null||k===undefined) return null; return k*1.852; }}
function fixText(d){{ if((d.fix_quality||0)<=0) return 'NO FIX'; return d.fix_type===3 ? '3D FIX' : (d.fix_type===2 ? '2D FIX' : 'FIX'); }}
async function upd(){{
 const r = await fetch('/api/position'); const d = await r.json();
 document.getElementById('fix').textContent = fixText(d);
 document.getElementById('lat').textContent = fmt(d.lat,6);
 document.getElementById('lon').textContent = fmt(d.lon,6);
 document.getElementById('alt').textContent = d.alt_m==null ? '--' : d.alt_m.toFixed(1)+' m';
 const sp = knotsToKmh(d.speed_knots);
 document.getElementById('speed').textContent = sp==null ? '--' : sp.toFixed(1)+' km/h';
 document.getElementById('course').textContent = d.course_deg==null ? '--' : d.course_deg.toFixed(1)+'°';
 document.getElementById('sats').textContent = d.sats_used ?? '--';
 document.getElementById('hdop').textContent = d.hdop==null ? '--' : d.hdop.toFixed(2);
 document.getElementById('utc').textContent = d.utc_text || '--';
 if(d.lat!=null && d.lon!=null){{ marker.setLatLng([d.lat,d.lon]); if(Array.isArray(d.track)) track.setLatLngs(d.track); if(firstFix){{ map.setView([d.lat,d.lon],17); firstFix=false; }} }}
}}
upd(); setInterval(upd, 1000);
</script>
</body>
</html>'''


class MapRequestHandler(BaseHTTPRequestHandler):
    mode = "online"

    def do_GET(self):
        if self.path == "/" or self.path.startswith("/index.html"):
            body = ONLINE_HTML if self.mode == "online" else offline_html()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(body.encode("utf-8"))
            return
        if self.path.startswith("/api/position"):
            with state_lock:
                payload = {
                    "lat": gnss_state.lat,
                    "lon": gnss_state.lon,
                    "alt_m": gnss_state.alt_m,
                    "speed_knots": gnss_state.speed_knots,
                    "course_deg": gnss_state.course_deg,
                    "utc_text": format_utc(gnss_state.utc_time, gnss_state.utc_date),
                    "fix_quality": gnss_state.fix_quality,
                    "fix_type": gnss_state.fix_type,
                    "sats_used": gnss_state.sats_used,
                    "gga_sats_used": gnss_state.gga_sats_used,
                    "hdop": gnss_state.hdop,
                    "track": list(gnss_state.track),
                }
            data = json.dumps(payload).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(data)
            return
        if self.mode == "offline" and self.path.startswith("/tiles/"):
            rel = unquote(self.path[len("/tiles/"):])
            tile_path = Path(TILES_ROOT) / rel
            if tile_path.is_file() and tile_path.suffix.lower() == ".png":
                self.send_response(200)
                self.send_header("Content-Type", "image/png")
                self.send_header("Cache-Control", "max-age=3600")
                self.end_headers()
                with open(tile_path, "rb") as f:
                    self.wfile.write(f.read())
                return
            self.send_response(404)
            self.end_headers()
            return
        self.send_response(404)
        self.end_headers()

    def log_message(self, fmt, *args):
        return


def make_handler(mode: str):
    class CustomHandler(MapRequestHandler):
        pass
    CustomHandler.mode = mode
    return CustomHandler


class MapServer(threading.Thread):
    def __init__(self, host: str, port: int, mode: str):
        super().__init__(daemon=True)
        self.host = host
        self.port = port
        self.mode = mode
        self.httpd = None
        self.error = ""

    def run(self):
        try:
            if self.mode == "offline":
                Path(TILES_ROOT).mkdir(parents=True, exist_ok=True)
            self.httpd = ThreadingHTTPServer((self.host, self.port), make_handler(self.mode))
            self.httpd.serve_forever()
        except Exception as e:
            self.error = str(e)

    def stop(self):
        if self.httpd:
            self.httpd.shutdown()


# ============================================================
# GUI
# ============================================================
class GNSSNavigatorApp:
    def __init__(self, root: tk.Tk, reader: GNSSReader, online_server: MapServer, offline_server: MapServer):
        self.root = root
        self.reader = reader
        self.online_server = online_server
        self.offline_server = offline_server
        self.sats_recording = False
        self.sats_log_path: Optional[Path] = None
        self.sats_wb: Optional[Workbook] = None
        self.sats_ws = None
        self.last_sats_log_time = 0.0
        self.last_sats_log_signature = None
        self.sats_log_rows_written = 0
        self.root.title(TITLE)
        self.root.configure(bg=BG)
        self.root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        self.root.minsize(680, 430)
        if FULLSCREEN:
            self.root.attributes("-fullscreen", True)
        self.root.bind("<Escape>", self.toggle_fullscreen)
        self.root.bind("q", self.quit_app)
        self.init_style()
        self.build_layout()
        self.root.after(REFRESH_MS, self.refresh_ui)

    def init_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background=PANEL, foreground=TEXT2, fieldbackground=PANEL, borderwidth=0, rowheight=22)
        style.configure("Treeview.Heading", background=PANEL2, foreground=TEXT, font=SMALL_BOLD_FONT)
        style.map("Treeview", background=[("selected", BLUE)], foreground=[("selected", "#ffffff")])
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(8, 4), font=TAB_FONT)
        style.configure("Dark.TCombobox", fieldbackground=PANEL2, background=PANEL2, foreground=TEXT, arrowcolor=TEXT)
        style.map("Dark.TCombobox", fieldbackground=[("readonly", PANEL2)], selectbackground=[("readonly", PANEL2)], selectforeground=[("readonly", TEXT)])

    def build_layout(self):
        self.header = tk.Frame(self.root, bg=PANEL, height=40)
        self.header.pack(fill="x", side="top")
        self.title_label = tk.Label(self.header, text="GNSS Navigator", fg=TEXT, bg=PANEL, font=HEADER_FONT)
        self.title_label.pack(side="left", padx=8, pady=6)
        self.status_label = tk.Label(self.header, text=f"UART: connecting... {SERIAL_PORT}", fg=MUTED, bg=PANEL, font=STATUS_FONT)
        self.status_label.pack(side="right", padx=8)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)
        self.tab_nav = tk.Frame(self.notebook, bg=BG)
        self.tab_sky = tk.Frame(self.notebook, bg=BG)
        self.tab_snr = tk.Frame(self.notebook, bg=BG)
        self.tab_diag = tk.Frame(self.notebook, bg=BG)
        self.tab_sats = tk.Frame(self.notebook, bg=BG)
        self.tab_const = tk.Frame(self.notebook, bg=BG)
        self.tab_map = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(self.tab_nav, text="NAV")
        self.notebook.add(self.tab_sky, text="SKY")
        self.notebook.add(self.tab_snr, text="SNR")
        self.notebook.add(self.tab_diag, text="DIAG")
        self.notebook.add(self.tab_sats, text="SATS")
        self.notebook.add(self.tab_const, text="Сузір'я")
        self.notebook.add(self.tab_map, text="MAP")
        self._last_sats_filter = "ALL"
        self.build_nav_tab()
        self.build_sky_tab()
        self.build_snr_tab()
        self.build_diag_tab()
        self.build_sats_tab()
        self.build_constellation_tab()
        self.build_map_tab()

    def toggle_fullscreen(self, event=None):
        current = bool(self.root.attributes("-fullscreen"))
        self.root.attributes("-fullscreen", not current)

    def quit_app(self, event=None):
        self.stop_sats_recording()
        self.reader.stop()
        self.online_server.stop()
        self.offline_server.stop()
        self.root.destroy()

    def snapshot_state(self):
        with state_lock:
            return GNSSState(
                lat=gnss_state.lat, lon=gnss_state.lon, alt_m=gnss_state.alt_m,
                speed_knots=gnss_state.speed_knots, course_deg=gnss_state.course_deg,
                utc_time=gnss_state.utc_time, utc_date=gnss_state.utc_date,
                fix_quality=gnss_state.fix_quality, fix_type=gnss_state.fix_type,
                sats_used=gnss_state.sats_used, pdop=gnss_state.pdop, hdop=gnss_state.hdop,
                vdop=gnss_state.vdop, mode_2d3d=gnss_state.mode_2d3d,
                last_sentence=gnss_state.last_sentence, last_update=gnss_state.last_update,
                satellites=dict(gnss_state.satellites), track=list(gnss_state.track),
                last_track_time=gnss_state.last_track_time,
                constellation_mode=gnss_state.constellation_mode,
                last_cmd=gnss_state.last_cmd,
                last_cmd_status=gnss_state.last_cmd_status,
            )

    def visible_sats(self):
        now = time.time()
        st = self.snapshot_state()
        sats = [s for s in st.satellites.values() if now - s.last_seen < 20.0]
        sats.sort(key=lambda s: (talker_name(s.talker), safe_int_sort(s.prn)))
        return sats

    def big_value(self, parent, title, value):
        card = tk.Frame(parent, bg=PANEL)
        card.pack(fill="x", padx=8, pady=5)
        tk.Label(card, text=title, bg=PANEL, fg="#95a7bf", font=NAV_TITLE_FONT, anchor="w").pack(fill="x")
        lbl = tk.Label(card, text=value, bg=PANEL, fg=TEXT, font=NAV_VALUE_FONT, anchor="w")
        lbl.pack(fill="x", pady=(2, 0))
        return lbl

    def build_nav_tab(self):
        top = tk.Frame(self.tab_nav, bg=BG)
        top.pack(fill="both", expand=True, padx=6, pady=6)
        left = tk.Frame(top, bg=PANEL)
        left.pack(side="left", fill="both", expand=True, padx=(0, 3))
        right = tk.Frame(top, bg=PANEL)
        right.pack(side="left", fill="both", expand=True, padx=(3, 0))
        self.fix_label = self.big_value(left, "FIX", "NO FIX")
        self.lat_label = self.big_value(left, "LAT", "--")
        self.lon_label = self.big_value(left, "LON", "--")
        self.alt_label = self.big_value(left, "ALT", "-- m")
        self.speed_label = self.big_value(right, "SPEED", "-- km/h")
        self.course_label = self.big_value(right, "COURSE", "--°")
        self.sats_used_label = self.big_value(right, "USED", "0")
        self.time_label = self.big_value(right, "UTC", "--:--:--")

    def build_sky_tab(self):
        wrap = tk.Frame(self.tab_sky, bg=BG)
        wrap.pack(fill="both", expand=True, padx=6, pady=6)
        self.sky_canvas = tk.Canvas(wrap, bg=PANEL, highlightthickness=0)
        self.sky_canvas.pack(fill="both", expand=True)

    def build_snr_tab(self):
        wrap = tk.Frame(self.tab_snr, bg=BG)
        wrap.pack(fill="both", expand=True, padx=6, pady=6)
        self.snr_canvas = tk.Canvas(wrap, bg=PANEL, highlightthickness=0)
        self.snr_canvas.pack(fill="both", expand=True)

    def build_diag_tab(self):
        container = tk.Frame(self.tab_diag, bg=BG)
        container.pack(fill="both", expand=True, padx=12, pady=6)
        self.diag_text = tk.Text(container, bg=PANEL, fg=TEXT2, insertbackground="#ffffff", font=MONO_FONT, relief="flat", wrap="word")
        self.diag_text.pack(fill="both", expand=True)

    def build_sats_tab(self):
        wrap = tk.Frame(self.tab_sats, bg=BG)
        wrap.pack(fill="both", expand=True, padx=6, pady=6)
        title_bar = tk.Frame(wrap, bg=PANEL)
        title_bar.pack(fill="x", pady=(0, 6))
        tk.Label(title_bar, text="Супутники: Elevation + SNR", bg=PANEL, fg=TEXT, font=SECTION_FONT).pack(side="left", padx=8, pady=6)
        tk.Label(title_bar, text="Фільтр:", bg=PANEL, fg=MUTED, font=SMALL_BOLD_FONT).pack(side="left", padx=(8, 4))
        self.sats_filter_var = tk.StringVar(value="ALL")
        self.sats_filter_box = ttk.Combobox(title_bar, textvariable=self.sats_filter_var, values=SATS_FILTER_VALUES, width=22, state="readonly", style="Dark.TCombobox")
        self.sats_filter_box.pack(side="left", padx=(0, 8))
        self.sats_filter_box.bind("<<ComboboxSelected>>", self.on_sats_filter_changed)
        tk.Label(title_bar, text="Запис:", bg=PANEL, fg=MUTED, font=SMALL_BOLD_FONT).pack(side="left", padx=(2, 4))
        self.sats_log_filter_var = tk.StringVar(value="ALL")
        self.sats_log_filter_box = ttk.Combobox(title_bar, textvariable=self.sats_log_filter_var, values=SATS_FILTER_VALUES, width=22, state="readonly", style="Dark.TCombobox")
        self.sats_log_filter_box.pack(side="left", padx=(0, 6))
        self.rec_btn = tk.Button(title_bar, text="Запис", bg="#2d5b3d", fg="white", font=SMALL_BOLD_FONT, relief="flat", command=self.toggle_sats_recording)
        self.rec_btn.pack(side="right", padx=4, pady=4)
        self.sats_log_label = tk.Label(title_bar, text="Лог: OFF", bg=PANEL, fg=MUTED, font=SMALL_FONT)
        self.sats_log_label.pack(side="right", padx=8)
        cols = ("sys", "prn", "elev", "azim", "snr", "used")
        self.sats_tree = ttk.Treeview(wrap, columns=cols, show="headings")
        for col, name, width in [("sys", "SYS", 90), ("prn", "PRN", 70), ("elev", "EL", 60), ("azim", "AZ", 70), ("snr", "SNR", 70), ("used", "USED", 70)]:
            self.sats_tree.heading(col, text=name)
            self.sats_tree.column(col, width=width, anchor="center")
        self.sats_tree.tag_configure("used", background="#203727")
        self.sats_tree.tag_configure("weak", foreground="#e39b3a")
        self.sats_tree.pack(fill="both", expand=True)

    def build_constellation_tab(self):
        wrap = tk.Frame(self.tab_const, bg=BG)
        wrap.pack(fill="both", expand=True, padx=10, pady=8)

        tk.Label(
            wrap,
            text="Обрати сузір'я L76K GPS HAT",
            bg=BG,
            fg=TEXT,
            font=SECTION_FONT
        ).pack(anchor="w", pady=(0, 6))

        text = (
            "Кнопки нижче змінюють робочий режим GNSS-модуля L76K. "
            "У UART передається тільки чиста PCAS-команда + CRLF.\n"
            "Після вибору режиму GPS / BeiDou / GLONASS або комбінованого режиму програма автоматично надсилає "
            "команду режиму і soft restart. Перевірка виконується у вкладці SATS."
        )
        tk.Label(
            wrap,
            text=text,
            bg=BG,
            fg=MUTED,
            font=SMALL_FONT,
            justify="left",
            wraplength=680
        ).pack(anchor="w", pady=(0, 8))

        grid = tk.Frame(wrap, bg=BG)
        grid.pack(fill="x")

        # Назва кнопки окремо, команда окремо. У serial.write() передається тільки cmd.
        self.make_const_button(grid, 0, 0, "GPS", "GPS + QZSS L1", "$PCAS04,1*18")
        self.make_const_button(grid, 0, 1, "BeiDou", "тільки BeiDou", "$PCAS04,2*1B")
        self.make_const_button(grid, 1, 0, "GLONASS", "тільки GLONASS", "$PCAS04,4*1D")
        self.make_const_button(grid, 1, 1, "GPS+BeiDou", "GPS + BeiDou", "$PCAS04,3*1A")
        self.make_const_button(grid, 2, 0, "GPS+GLONASS", "GPS + GLONASS", "$PCAS04,5*1C")
        self.make_const_button(grid, 2, 1, "BeiDou+GLONASS", "BeiDou + GLONASS", "$PCAS04,6*1F")
        self.make_const_button(grid, 3, 0, "GPS+BeiDou+GLONASS", "усі основні", "$PCAS04,7*1E")
        self.make_cmd_button(grid, 3, 1, "Перезапуск GNSS", "soft restart", "$PCAS10,2*1E", "Restart")

        self.const_status = tk.Text(
            wrap,
            height=9,
            bg=PANEL,
            fg=TEXT2,
            insertbackground="#ffffff",
            font=MONO_FONT,
            relief="flat",
            wrap="word"
        )
        self.const_status.pack(fill="both", expand=True, pady=(10, 0))
        self.const_status.insert("1.0", "Статус: очікування команди\n")
        self.const_status.insert(tk.END, f"Порт: {SERIAL_PORT}    Baudrate: {BAUDRATE}\n")
        self.const_status.insert(tk.END, "У UART буде передано тільки рядок $PCAS... + CRLF.\n")

    def make_const_button(self, parent, row, col, title, subtitle, cmd):
        btn_text = f"{title}\n{subtitle}"
        btn = tk.Button(
            parent,
            text=btn_text,
            bg=PANEL2,
            fg=TEXT,
            activebackground=BLUE,
            activeforeground="white",
            font=SMALL_BOLD_FONT,
            justify="center",
            relief="flat",
            command=lambda c=cmd, n=title: self.set_constellation(n, c)
        )
        btn.grid(row=row, column=col, sticky="nsew", padx=5, pady=5, ipady=10)
        parent.grid_columnconfigure(col, weight=1)

    def make_cmd_button(self, parent, row, col, title, subtitle, cmd, mode_name):
        btn_text = f"{title}\n{subtitle}"
        btn = tk.Button(
            parent,
            text=btn_text,
            bg="#3a4656",
            fg=TEXT,
            activebackground=BLUE,
            activeforeground="white",
            font=SMALL_BOLD_FONT,
            justify="center",
            relief="flat",
            command=lambda c=cmd, n=mode_name: self.set_constellation(n, c)
        )
        btn.grid(row=row, column=col, sticky="nsew", padx=5, pady=5, ipady=10)
        parent.grid_columnconfigure(col, weight=1)

    def set_constellation(self, name: str, cmd: str):
        # У serial TX передаються тільки PCAS-рядки. Назви кнопок не передаються.
        restart_cmd = "$PCAS10,2*1E"
        if cmd.startswith("$PCAS04"):
            tx_commands = [cmd, restart_cmd]
        else:
            tx_commands = [cmd]

        ok = self.reader.queue_commands(tx_commands)
        ts = datetime.now().strftime("%H:%M:%S")
        raw_lines = [repr((c + "\r\n").encode("ascii")) for c in tx_commands]

        if ok:
            status = f"OK {ts}: команди поставлені в UART-чергу"
        else:
            err = self.reader.last_error or "UART ще не відкритий або порт зайнятий"
            status = f"ERROR {ts}: {err}"

        active_filter = mode_to_sats_filter(name)
        if active_filter is not None:
            # Після перемикання модуля одразу переводимо SATS у відповідний фільтр
            # і очищаємо старі супутники інших сузір'їв з попереднього режиму.
            self.sats_filter_var.set(active_filter)
            self._last_sats_filter = active_filter
            # Повне очищення кешу: після перемикання режиму старі GSV/GSA
            # не мають впливати на SATS/NAV. Нові дані прийдуть вже після restart.
            clear_satellite_cache()

        with state_lock:
            gnss_state.constellation_mode = name
            gnss_state.last_cmd = " ; ".join(tx_commands)
            gnss_state.last_cmd_status = status

        self.const_status.delete("1.0", tk.END)
        self.const_status.insert(tk.END, status + "\n")
        self.const_status.insert(tk.END, f"Режим у GUI: {name}\n")
        self.const_status.insert(tk.END, "TX UART, тільки ці байти:\n")
        for raw in raw_lines:
            self.const_status.insert(tk.END, f"  {raw}\n")
        self.const_status.insert(tk.END, "\nПояснення:\n")
        self.const_status.insert(tk.END, "- Для всіх режимів PCAS04 автоматично додається soft restart.\n")
        self.const_status.insert(tk.END, "- Назви кнопок у порт НЕ передаються.\n")
        self.const_status.insert(tk.END, "- SATS автоматично перемикає фільтр на вибраний режим, включно з GLONASS та комбінованими режимами.\n")
        self.const_status.insert(tk.END, "- Після перемикання очищується весь кеш GSV/GSA; USED береться тільки з нових GSA.\n\n")
        self.const_status.insert(tk.END, f"Порт: {SERIAL_PORT}\n")
        self.const_status.insert(tk.END, f"Baudrate: {BAUDRATE}\n")

    def build_map_tab(self):
        wrap = tk.Frame(self.tab_map, bg=BG)
        wrap.pack(fill="both", expand=True, padx=8, pady=8)
        top = tk.Frame(wrap, bg=BG)
        top.pack(fill="x")
        tk.Button(top, text="Відкрити Online Map", bg=PANEL2, fg=TEXT, activebackground=BLUE, relief="flat", font=SMALL_BOLD_FONT, command=self.open_online_map).pack(side="left", padx=(0, 6), ipady=5)
        tk.Button(top, text="Відкрити Offline Map", bg=PANEL2, fg=TEXT, activebackground=BLUE, relief="flat", font=SMALL_BOLD_FONT, command=self.open_offline_map).pack(side="left", padx=(0, 6), ipady=5)
        self.map_info = tk.Text(wrap, bg=PANEL, fg=TEXT2, insertbackground="#ffffff", font=MONO_FONT, relief="flat", wrap="word")
        self.map_info.pack(fill="both", expand=True, pady=(8, 0))

    def open_online_map(self):
        webbrowser.open(f"http://127.0.0.1:{MAP_ONLINE_PORT}")

    def open_offline_map(self):
        webbrowser.open(f"http://127.0.0.1:{MAP_OFFLINE_PORT}")

    def on_sats_filter_changed(self, event=None):
        self._last_sats_filter = self.sats_filter_var.get().strip()
        st = self.snapshot_state()
        self.draw_sats(st, self.visible_sats())

    def filter_sats(self, sats, filter_name: str):
        return [sat for sat in sats if satellite_allowed_by_filter(sat, filter_name)]

    def sats_rows_for_display(self, st: GNSSState, sats):
        filtered = self.filter_sats(sats, self.sats_filter_var.get())
        rows = []
        for sat in filtered:
            tags = []
            if sat.used:
                tags.append("used")
            if sat.snr is not None and sat.snr < 25:
                tags.append("weak")
            rows.append({
                "values": (
                    talker_name(sat.talker), sat.prn,
                    "--" if sat.elev is None else sat.elev,
                    "--" if sat.azim is None else sat.azim,
                    "--" if sat.snr is None else sat.snr,
                    "YES" if sat.used else "",
                ),
                "tags": tuple(tags),
            })
        return rows

    def refresh_ui(self):
        st = self.snapshot_state()
        sats = self.visible_sats()
        age = time.time() - st.last_update if st.last_update else 9999
        uart_state = "OK" if age < 3.0 else "NO DATA"
        if self.reader.last_error:
            uart_state = "ERROR"
        self.status_label.config(text=f"UART {uart_state}: {SERIAL_PORT} | {BAUDRATE} | {st.constellation_mode}")
        self.draw_nav(st)
        self.draw_sky(sats)
        self.draw_snr(sats)
        self.draw_diag(st, sats)
        self.draw_sats(st, sats)
        self.draw_map_info(st)
        self.write_sats_log_if_needed(st, sats)
        self.root.after(REFRESH_MS, self.refresh_ui)

    def draw_nav(self, st):
        self.fix_label.config(text=fix_text(st), fg=GREEN if st.fix_quality > 0 else RED)
        self.lat_label.config(text="--" if st.lat is None else f"{st.lat:.7f}")
        self.lon_label.config(text="--" if st.lon is None else f"{st.lon:.7f}")
        self.alt_label.config(text=f"{fmt_float(st.alt_m, 1)} m")
        sp = knots_to_kmh(st.speed_knots)
        self.speed_label.config(text=f"{fmt_float(sp, 1)} km/h")
        self.course_label.config(text=f"{fmt_float(st.course_deg, 1)}°")
        try:
            sats_now = self.visible_sats()
            visible_n, used_n = count_visible_used(sats_now, self.sats_filter_var.get())
            self.sats_used_label.config(text=f"{used_n}/{visible_n}")
        except Exception:
            self.sats_used_label.config(text=str(st.sats_used))
        self.time_label.config(text=format_utc(st.utc_time, st.utc_date))

    def draw_sky(self, sats):
        c = self.sky_canvas
        c.delete("all")
        w = max(c.winfo_width(), 200)
        h = max(c.winfo_height(), 200)
        cx, cy = w / 2, h / 2
        r = min(w, h) * 0.42
        c.create_oval(cx - r, cy - r, cx + r, cy + r, outline="#506070", width=2)
        c.create_oval(cx - r / 2, cy - r / 2, cx + r / 2, cy + r / 2, outline="#344252")
        c.create_line(cx - r, cy, cx + r, cy, fill="#344252")
        c.create_line(cx, cy - r, cx, cy + r, fill="#344252")
        c.create_text(cx, cy - r - 10, text="N", fill=TEXT, font=SMALL_BOLD_FONT)
        c.create_text(cx, cy + r + 10, text="S", fill=TEXT, font=SMALL_BOLD_FONT)
        c.create_text(cx + r + 10, cy, text="E", fill=TEXT, font=SMALL_BOLD_FONT)
        c.create_text(cx - r - 10, cy, text="W", fill=TEXT, font=SMALL_BOLD_FONT)
        for sat in sats:
            if sat.elev is None or sat.azim is None:
                continue
            rr = r * (90 - max(0, min(90, sat.elev))) / 90.0
            a = math.radians(sat.azim - 90)
            x = cx + rr * math.cos(a)
            y = cy + rr * math.sin(a)
            col = snr_to_color(sat.snr)
            c.create_oval(x - 9, y - 9, x + 9, y + 9, fill=col, outline="")
            c.create_text(x, y, text=sat.prn, fill=BG, font=("DejaVu Sans", 7, "bold"))
            c.create_text(x, y + 15, text=sat.talker, fill=MUTED, font=("DejaVu Sans", 7))

    def draw_snr(self, sats):
        c = self.snr_canvas
        c.delete("all")
        w = max(c.winfo_width(), 200)
        h = max(c.winfo_height(), 200)
        left, top, bottom = 35, 20, 38
        plot_w = w - left - 10
        plot_h = h - top - bottom
        c.create_text(10, 10, anchor="nw", text="SNR dB-Hz", fill=TEXT, font=SMALL_BOLD_FONT)
        if not sats:
            c.create_text(w / 2, h / 2, fill="#95a7bf", font=("DejaVu Sans", 11), text="Немає даних SNR")
            return
        gap = 7
        bar_w = max(12, (plot_w - gap * (len(sats) + 1)) / max(1, len(sats)))
        x = left + gap
        for sat in sats:
            snr = max(0, min(60, sat.snr or 0))
            bh = (snr / 60.0) * plot_h
            y0 = top + plot_h - bh
            c.create_rectangle(x, y0, x + bar_w, top + plot_h, fill=snr_to_color(snr), outline="")
            c.create_text(x + bar_w / 2, y0 - 10, fill=TEXT2, font=("DejaVu Sans", 9, "bold"), text=str(snr))
            c.create_text(x + bar_w / 2, top + plot_h + 14, fill=TEXT2, font=SMALL_FONT, text=sat.prn)
            c.create_text(x + bar_w / 2, top + plot_h + 28, fill="#89a3bf", font=("DejaVu Sans", 8), text=sat.talker)
            x += bar_w + gap

    def draw_diag(self, st: GNSSState, sats):
        by_sys = defaultdict(int)
        for s in sats:
            by_sys[talker_name(s.talker)] += 1
        lines = []
        lines.append("ДІАГНОСТИКА GNSS\n")
        lines.append(f"Порт: {SERIAL_PORT}")
        lines.append("Вибір USB/UART робиться через GNSS_SERIAL_PORT")
        lines.append(f"UART baudrate: {BAUDRATE}")
        lines.append(f"Сузір'я режим: {st.constellation_mode}")
        lines.append(f"Остання команда: {st.last_cmd or '--'}")
        lines.append(f"Статус команди: {st.last_cmd_status or '--'}")
        lines.append(f"Fix quality: {st.fix_quality}")
        lines.append(f"Fix type: {st.fix_type}")
        lines.append(f"Mode 2D/3D: {st.mode_2d3d}")
        visible_f, used_f = count_visible_used(sats, self.sats_filter_var.get())
        lines.append(f"Satellites used from GSA, filtered: {used_f}")
        lines.append(f"Visible satellites from GSV, filtered: {visible_f}")
        lines.append(f"Satellites used from GSA, all cached: {st.sats_used}")
        lines.append(f"Raw GGA satellites field: {st.gga_sats_used}")
        lines.append(f"Visible satellites all cached: {len(sats)}")
        lines.append(f"PDOP: {fmt_float(st.pdop)}")
        lines.append(f"HDOP: {fmt_float(st.hdop)}")
        lines.append(f"VDOP: {fmt_float(st.vdop)}")
        lines.append(f"Altitude: {fmt_float(st.alt_m, 1)} m")
        lines.append(f"Speed: {fmt_float(knots_to_kmh(st.speed_knots), 1)} km/h")
        lines.append(f"Course: {fmt_float(st.course_deg, 1)}°")
        lines.append("")
        lines.append("По системах:")
        for name in sorted(by_sys):
            lines.append(f"  {name}: {by_sys[name]}")
        lines.append("")
        lines.append("Останній NMEA:")
        lines.append(st.last_sentence or "--")
        if self.reader.last_error:
            lines.append("")
            lines.append("UART error:")
            lines.append(self.reader.last_error)
        if self.online_server.error or self.offline_server.error:
            lines.append("")
            lines.append("Map server error:")
            if self.online_server.error:
                lines.append(f"Online: {self.online_server.error}")
            if self.offline_server.error:
                lines.append(f"Offline: {self.offline_server.error}")
        self.diag_text.delete("1.0", tk.END)
        self.diag_text.insert("1.0", "\n".join(lines))

    def draw_sats(self, st: GNSSState, sats):
        for item in self.sats_tree.get_children():
            self.sats_tree.delete(item)
        rows = self.sats_rows_for_display(st, sats)
        for row in rows:
            self.sats_tree.insert("", "end", values=row["values"], tags=row["tags"])

    def draw_map_info(self, st: GNSSState):
        lines = []
        lines.append("MAP CONTROL\n")
        lines.append(f"Online map URL : http://127.0.0.1:{MAP_ONLINE_PORT}")
        lines.append(f"Offline map URL: http://127.0.0.1:{MAP_OFFLINE_PORT}")
        lines.append(f"Offline tiles   : {TILES_ROOT}")
        lines.append("")
        lines.append("Що роблять кнопки:")
        lines.append("  - Відкрити Online Map  -> карта OpenStreetMap через інтернет")
        lines.append("  - Відкрити Offline Map -> локальні PNG тайли з ~/gnss_tiles")
        lines.append("")
        lines.append("Поточна позиція:")
        lines.append(f"  Lat : {self.format_latlon(st.lat, True)}")
        lines.append(f"  Lon : {self.format_latlon(st.lon, False)}")
        lines.append(f"  Fix : {fix_text(st)}")
        lines.append(f"  UTC : {format_utc(st.utc_time, st.utc_date)}")
        lines.append("")
        lines.append("Формат локальних тайлів:")
        lines.append("  ~/gnss_tiles/z/x/y.png")
        lines.append("  приклад: ~/gnss_tiles/12/2203/1345.png")
        lines.append("")
        lines.append("Excel лог SATS:")
        lines.append(f"  Каталог: {LOGS_ROOT}")
        lines.append("  Фільтр 'Запис' задає, які сузір'я записувати в .xlsx")
        lines.append("")
        lines.append("Трек:")
        lines.append(f"  Track points: {len(st.track)}")
        lines.append("  Кожна нова координата додається в polyline на карті.")
        self.map_info.delete("1.0", tk.END)
        self.map_info.insert("1.0", "\n".join(lines))

    def format_latlon(self, v: Optional[float], is_lat: bool) -> str:
        if v is None:
            return "--"
        hemi = "N" if is_lat and v >= 0 else "S" if is_lat else "E" if v >= 0 else "W"
        return f"{abs(v):.7f}° {hemi}"

    def toggle_sats_recording(self):
        if self.sats_recording:
            self.stop_sats_recording()
        else:
            self.start_sats_recording()

    def start_sats_recording(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Excel", "Потрібен пакет openpyxl: sudo apt install python3-openpyxl")
            return
        LOGS_ROOT.mkdir(parents=True, exist_ok=True)
        log_filter = self.sats_log_filter_var.get().replace("+", "plus")
        self.sats_log_path = LOGS_ROOT / f"sats_{log_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.sats_wb = Workbook()
        self.sats_ws = self.sats_wb.active
        self.sats_ws.title = "SATS"
        headers = ["time", "log_filter", "sys", "talker", "prn", "elev", "azim", "snr", "used", "lat", "lon", "alt_m", "fix", "hdop", "pdop", "utc"]
        self.sats_ws.append(headers)
        for cell in self.sats_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="243140")
        self.sats_log_rows_written = 0
        self.last_sats_log_time = 0.0
        self.last_sats_log_signature = None
        self.sats_recording = True
        self.rec_btn.config(text="Стоп", bg="#7a3232")
        self.sats_log_label.config(text=f"Лог: ON {self.sats_log_filter_var.get()}")

    def stop_sats_recording(self):
        if self.sats_recording and self.sats_wb and self.sats_log_path:
            try:
                for idx, width in enumerate([20, 12, 12, 8, 8, 8, 8, 8, 8, 13, 13, 10, 10, 8, 8, 16], start=1):
                    self.sats_ws.column_dimensions[get_column_letter(idx)].width = width
                self.sats_wb.save(self.sats_log_path)
            except Exception as e:
                print("Excel save error:", e)
        self.sats_recording = False
        self.sats_wb = None
        self.sats_ws = None
        if hasattr(self, "rec_btn"):
            self.rec_btn.config(text="Запис", bg="#2d5b3d")
        if hasattr(self, "sats_log_label"):
            self.sats_log_label.config(text="Лог: OFF")

    def write_sats_log_if_needed(self, st: GNSSState, sats):
        if not self.sats_recording or not self.sats_ws:
            return
        now = time.time()
        if now - self.last_sats_log_time < SATS_LOG_MIN_INTERVAL_SEC:
            return
        filtered = self.filter_sats(sats, self.sats_log_filter_var.get())
        signature = tuple((s.talker, s.prn, s.elev, s.azim, s.snr, s.used) for s in filtered)
        if signature == self.last_sats_log_signature:
            return
        self.last_sats_log_signature = signature
        self.last_sats_log_time = now
        t = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for sat in filtered:
            self.sats_ws.append([
                t, self.sats_log_filter_var.get(), talker_name(sat.talker), sat.talker, sat.prn,
                sat.elev, sat.azim, sat.snr, "YES" if sat.used else "",
                st.lat, st.lon, st.alt_m, fix_text(st), st.hdop, st.pdop,
                format_utc(st.utc_time, st.utc_date),
            ])
            self.sats_log_rows_written += 1
        self.sats_log_label.config(text=f"Лог: ON {self.sats_log_filter_var.get()} {self.sats_log_rows_written}")


# ============================================================
# main
# ============================================================
def main():
    parser = NMEAParser()
    reader = GNSSReader(parser, SERIAL_PORT, BAUDRATE)
    online_server = MapServer("0.0.0.0", MAP_ONLINE_PORT, "online")
    offline_server = MapServer("0.0.0.0", MAP_OFFLINE_PORT, "offline")
    reader.start()
    online_server.start()
    offline_server.start()
    root = tk.Tk()
    app = GNSSNavigatorApp(root, reader, online_server, offline_server)
    root.protocol("WM_DELETE_WINDOW", app.quit_app)
    root.mainloop()


if __name__ == "__main__":
    main()
